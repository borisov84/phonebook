import openpyxl as op
from lxml import etree

filename = 'справка.xlsx'
wb = op.load_workbook(filename)
sheet_ranges = wb['Справочник']
max_row = sheet_ranges.max_row
filials = ['Управление организации восстановления основных фондов (УОВОФ)',
           'Югорское Управление материально-технического снабжения и комплектации (ЮУМТСиК)',
           'Югорское Управление технологическим транспортом \nи специальной техники (ЮУТТиСТ)',
           'Югорское Управление \nаварийно-восстановительных работ (ЮУАВР)',
           'Инженерно-технический центр (ИТЦ)',
           'Управление по эксплуатации зданий и сооружений (УЭЗиС)',
           'Управление связи', 'Санаторий-профилакторий', 'Культурно-спортивный комплекс "НОРД"',
           'Ямбургское ЛПУМГ    ( код 778 )', 'Ныдинское ЛПУМГ          ( код 778 )',
           'Ново-Уренгойское ЛПУМГ   ( код 778 )', 'Пангодинское ЛПУМГ            ( код 778 )',
           'Правохеттинское ЛПУМГ  (код 778)', 'Надымское ЛПУМГ (код 778)',
           'Надымское УТТиСТ           ( код 778 )',
           'Надымское управление \nаварийно-восстановительных работ (код 778)',
           'Ягельное ЛПУМГ           ( код 778 )', 'Приозерное ЛПУМГ  (код 778)',
           'Лонг-Юганское ЛПУМГ           ( код 778 )', 'Сосновское ЛПУМГ  ', 'Сорумское ЛПУМГ  ',
           'Верхнеказымское ЛПУМГ', 'Казымское ЛПУМГ', 'Белоярское УТТиСТ  ',
           'Белоярское управление аварийно-восстановительных работ (БУАВР)', 'Бобровское ЛПУМГ',
           'Октябрьское ЛПУМГ', 'Перегребненское  ЛПУМГ', 'Учебно-производственный центр', 'Пунгинское ЛПУМГ   ',
           'Сосьвинское ЛПУМГ', 'Уральское ЛПУМГ', 'Таежное ЛПУМГ', 'Комсомольское ЛПУМГ', 'Пелымское ЛПУМГ  ',
           'Ивдельское ЛПУМГ', 'Карпинское ЛПУМГ', 'Краснотурьинское ЛПУМГ', 'Нижнетуринское ЛПУМГ']

filials_normal = ['УОВОФ', 'ЮУМТСиК', 'ЮУТТиСТ', 'ЮУАВР', 'ИТЦ', 'УЭЗиС', 'Управление связи', 'Санаторий-профилакторий',
                  'КСК "НОРД"', 'Ямбургское ЛПУМГ', 'Ныдинское ЛПУМГ', 'Ново-Уренгойское ЛПУМГ', 'Пангодинское ЛПУМГ',
                  'Правохеттинское ЛПУМГ', 'Надымское ЛПУМГ', 'Надымское УТТиСТ', 'Надымское УАВР', 'Ягельное ЛПУМГ',
                  'Приозерное ЛПУМГ', 'Лонг-Юганское ЛПУМГ', 'Сосновское ЛПУМГ  ', 'Сорумское ЛПУМГ  ',
                  'Верхнеказымское ЛПУМГ', 'Казымское ЛПУМГ', 'Белоярское УТТиСТ', 'Белоярское УАВР',
                  'Бобровское ЛПУМГ', 'Октябрьское ЛПУМГ', 'Перегребненское  ЛПУМГ', 'УПЦ', 'Пунгинское ЛПУМГ',
                  'Сосьвинское ЛПУМГ', 'Уральское ЛПУМГ', 'Таежное ЛПУМГ', 'Комсомольское ЛПУМГ', 'Пелымское ЛПУМГ',
                  'Ивдельское ЛПУМГ', 'Карпинское ЛПУМГ', 'Краснотурьинское ЛПУМГ', 'Нижнетуринское ЛПУМГ']

filial_rows = []
filial_ranges = []
subscribers = []
with open('allowed_deps.txt', 'r', encoding='utf-8') as al_deps:
    allowed_deps = al_deps.read().splitlines()
al_deps.close()
# print(f'Филиалов: {len(allowed_deps)}')


def get_departments_rows():
    for num in sheet_ranges.iter_rows(min_row=1, max_col=1, max_row=max_row):
        for filial in filials:
            if num[0].value == filial:
                filial_rows.append(num[0].row)
    filial_rows.append(max_row)
    for index, rng in enumerate(filial_rows):
        if index + 1 < len(filial_rows):
            filial_ranges.append([rng, filial_rows[index + 1] - 1])
    # print(f'Найдено {len(filial_ranges)} диапазонов для {len(filials)} филиалов')
    # print(filial_rows)


def get_subs():
    filial = ""
    for num in sheet_ranges.iter_rows(min_row=9, max_col=3, max_row=max_row):
        # 0 - Фамилия, 1 - должность, 2 - номер
        if num[2].value and num[0].value and (len(str(num[2].value).strip()) == 7 or
                                              ((len(str(num[2].value).strip()) == 6) and str(
                                                  num[2].value[2:3] == "-"))):
            # print(f'{num[2].row} {num[2].value}')
            if sheet_ranges.cell(row=num[0].row + 1, column=1).value and num[0].value.strip().find(" ") == -1:
                io_cell = sheet_ranges.cell(row=num[0].row + 1, column=1).value
                family = num[0].value
                name = io_cell[0:1] + "."
                otch = io_cell[io_cell.find(" ") + 1:io_cell.find(" ") + 2] + "."
                for ind, filial_range in enumerate(filial_ranges):
                    if filial_range[0] < num[0].row < filial_range[1]:
                        filial = filials_normal[ind]
                full_fio = family.strip() + " " + name + otch
                subscribers.append((filial, full_fio, num[2].value.strip()))


def export_phonebook_yealink():
    root = etree.Element("YealinkIPPhoneBook")
    title = etree.SubElement(root, "Title")
    title.text = "Yealink"
    departments = []
    for subscriber in subscribers:
        if subscriber[0] in allowed_deps:
            if subscriber[0] not in departments:
                departments.append(subscriber[0])
                menu = etree.SubElement(root, "Menu", Name=subscriber[0].replace('\n', ""))
                for i in subscribers:
                    if i[0] == subscriber[0]:
                        etree.SubElement(menu, "Unit", Name=i[1], Phone1=i[2].replace("-", ""))
    et = etree.ElementTree(root)
    et.write('yug_fil.xml', pretty_print=True, encoding='utf-8', xml_declaration=True)

def export_phonebook_eltex():
    root = etree.Element('EltexIPPhoneDirectory')
    title = etree.SubElement(root, 'Title')
    title.text = 'EltexPhones'
    prompt = etree.SubElement(root, 'Prompt')
    prompt.text = 'Prompt'
    gr_list = etree.SubElement(root, 'Grouplist')
    deps = []
    for subscriber in subscribers:
        if subscriber[0] in allowed_deps:
            if subscriber[0] not in deps:
                deps.append(subscriber[0])
                etree.SubElement(gr_list, 'Group', name=subscriber[0].replace('\n', ''))
    for subscriber in subscribers:
        if subscriber[0] in allowed_deps:
            dir_entry = etree.SubElement(root, 'DirectoryEntry')
            name = etree.SubElement(dir_entry, "Name")
            name.text = subscriber[1]
            telephone = etree.SubElement(dir_entry, 'Telephone')
            telephone.text = subscriber[2].replace('-', '')
            grp = etree.SubElement(dir_entry, 'Group')
            grp.text = subscriber[0]
    et = etree.ElementTree(root)
    et.write('yug_fil_eltex.xml', pretty_print=True, encoding = 'utf-8', xml_declaration=True)


if __name__ == '__main__':
    get_departments_rows()
    get_subs()
    export_phonebook_yealink()
    export_phonebook_eltex()
