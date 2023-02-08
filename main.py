import openpyxl as op
from lxml import etree


def get_departments_rows():
    for num in sheet_ranges.iter_rows(min_row=1, max_col=1, max_row=max_row):
        for filial in filials:
            if num[0].value == filial:
                filial_rows.append(num[0].row)
    filial_rows.append(max_row)
    for index, rng in enumerate(filial_rows):
        if index + 1 < len(filial_rows):
            filial_ranges.append([rng, filial_rows[index + 1] - 1])
    print(f'Найдено {len(filial_ranges)} диапазонов для {len(filials)} филиалов')
    # print(filial_rows)


def get_subs():
    filial = ""
    for num in sheet_ranges.iter_rows(min_row=9, max_col=3, max_row=max_row):
        # print(num[2].value)
        # 0 - Фамилия, 1 - должность, 2 - номер
        if num[2].value and num[0].value and len(str(num[2].value)) == 7:
            # print(f'{num[2].row} {num[2].value}')
            if sheet_ranges.cell(row=num[0].row + 1, column=1).value and sheet_ranges.cell(row=num[0].row, column=1).font.bold == True:
                io_cell = sheet_ranges.cell(row=num[0].row + 1, column=1).value
                family = num[0].value
                name = io_cell[0:1] + "."
                otch = io_cell[io_cell.find(" ") + 1:io_cell.find(" ") + 2] + ". "
                for ind, filial_range in enumerate(filial_ranges):
                    if filial_range[0] < num[0].row < filial_range[1]:
                        filial = filials[ind]
                full_fio = family.strip() + " " + name + otch
                # print(f'{filial} {num[0].value} {name}{otch} {num[2].value}')
                subscribers.append((filial, full_fio, num[2].value))
    # print(subscribers)


def test():
    fil = []
    for subscriber in subscribers:
        if subscriber[0] not in fil:
            # print(subscriber[0])
            fil.append(subscriber[0])
    print(fil)


def export_phonebook():
    root = etree.Element("YealinkIPPhoneBook")
    title = etree.SubElement(root, "Title")
    title.text = "Yealink"
    departments = []
    for subscriber in subscribers:
        if subscriber[0] not in departments:
            departments.append(subscriber[0])
            menu = etree.SubElement(root, "Menu", Name=subscriber[0].replace('\n', ""))
            for i in subscribers:
                if i[0] == subscriber[0]:
                    etree.SubElement(menu, "Item", Name=i[1], Phone1=i[2].replace("-", ""))

    et = etree.ElementTree(root)
    et.write('output_yealink.xml', pretty_print=True, encoding='utf-8', xml_declaration=True)

    # with open("phnbk.xml", "w") as xml_file:
    #     xml_file.write(etree.tostring(root, pretty_print=True))
    # print(etree.tostring(root, pretty_print=True))


if __name__ == '__main__':
    filename = '../справка.xlsx'
    wb = op.load_workbook(filename)
    sheet_ranges = wb['Справочник']
    max_row = sheet_ranges.max_row
    filials = ['Управление организации восстановления основных фондов (УОВОФ)',
               'Югорское Управление материально-технического снабжения и комплектации (ЮУМТСиК)',
               'Югорское Управление технологическим транспортом \nи специальной техники (ЮУТТиСТ)',
               'Югорское Управление \nаварийно-восстановительных работ (ЮУАВР)',
               'Инженерно-технический центр (ИТЦ)',
               'Управление по эксплуатации зданий и сооружений (УЭЗиС)',
               'Управление связи', 'Санаторий-профилакторий', 'Культурно-спортивный комплекс "НОРД"',
               'Ямбургское ЛПУМГ    ( код 778 )', 'Ныдинское ЛПУМГ          ( код 778 )',
               'Ново-Уренгойское ЛПУМГ   ( код 778 )', 'Пангодинское ЛПУМГ            ( код 778 )',
               'Правохеттинское ЛПУМГ  (код 778)', 'Надымское ЛПУМГ (код 778)',
               'Надымское УТТиСТ           ( код 778 )',
               'Надымское управление \nаварийно-восстановительных работ (код 778)',
               'Ягельное ЛПУМГ           ( код 778 )', 'Приозерное ЛПУМГ  (код 778)',
               'Лонг-Юганское ЛПУМГ           ( код 778 )', 'Сосновское ЛПУМГ  ', 'Сорумское ЛПУМГ  ',
               'Верхнеказымское ЛПУМГ', 'Казымское ЛПУМГ', 'Белоярское УТТиСТ  ',
               'Белоярское управление аварийно-восстановительных работ (БУАВР)', 'Бобровское ЛПУМГ',
               'Октябрьское ЛПУМГ', 'Перегребненское  ЛПУМГ', 'Учебно-производственный центр', 'Пунгинское ЛПУМГ   ',
               'Сосьвинское ЛПУМГ', 'Уральское ЛПУМГ', 'Таежное ЛПУМГ', 'Комсомольское ЛПУМГ', 'Пелымское ЛПУМГ  ',
               'Ивдельское ЛПУМГ', 'Карпинское ЛПУМГ', 'Краснотурьинское ЛПУМГ', 'Нижнетуринское ЛПУМГ']
    filial_rows = []
    filial_ranges = []
    subscribers = []
    get_departments_rows()
    get_subs()
    # print(subscribers)
    # test()
    export_phonebook()
