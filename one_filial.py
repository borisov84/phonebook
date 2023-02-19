import openpyxl as op
from lxml import etree
from sys import argv


path, deps, deps_normal = argv
filename = 'справка.xlsx'
wb = op.load_workbook(filename)
sheet_ranges = wb['Справочник']
max_row = sheet_ranges.max_row
departments = []
filial_rows = []
filial_ranges = []
filial_range = []
subscribers = []
filial_normal = []
output_file = deps.split('_')[0] + '.xml'


def get_departments_names():
    with open(deps, 'r', encoding='utf-8') as dep_names:
        global departments
        departments = dep_names.read().splitlines()
    dep_names.close()
    for num in sheet_ranges.iter_rows(min_row=1, max_col=1, max_row=max_row):
        if num[0].value == departments[0]:
            departments[0] = num[0].row
        if num[0].value == departments[1]:
            departments[1] = num[0].row
    # print(departments)


def get_normal_names():
    global filial_normal
    with open(deps_normal, 'r', encoding='utf-8') as us_nor:
        filial_normal = us_nor.read().splitlines()
    us_nor.close()


def get_filial_rows():
    for num in sheet_ranges.iter_rows(min_row=int(departments[0]), max_col=1, max_row=int(departments[1])):
        for filial in departments[2:]:
            if num[0].value == filial:
                filial_rows.append(num[0].row)
    filial_rows.append(int(departments[1]))
    for index, rng in enumerate(filial_rows):
        if index + 1 < len(filial_rows):
            filial_ranges.append([rng, filial_rows[index + 1] - 1])
    print(f'Найдено {len(filial_ranges)} диапазонов для {len(departments) - 2} филиалов')
    # print(filial_ranges)


def get_subs():
    filial = ""
    global filial_range
    global departments
    for num in sheet_ranges.iter_rows(min_row=departments[0], max_col=3, max_row=departments[1]):
        # print(num[2].value)
        # 0 - Фамилия, 1 - должность, 2 - номер
        # print(f'{num[2].value}  {type(num[2].value)}')
        if num[2].value and num[0].value and (len(str(num[2].value).strip()) == 7 or
                                              ((len(str(num[2].value).strip()) == 6) and str(
                                                  num[2].value[2:3] == "-"))):
            # print(f'{num[2].row} {num[2].value}')
            if sheet_ranges.cell(row=num[0].row + 1, column=1).value and num[0].value.strip().find(" ") == -1:
                io_cell = sheet_ranges.cell(row=num[0].row + 1, column=1).value
                family = num[0].value
                name = io_cell[0:1] + "."
                otch = io_cell[io_cell.find(" ") + 1:io_cell.find(" ") + 2] + "."
                for ind, filial_range in enumerate(filial_ranges):
                    if filial_range[0] < num[0].row < filial_range[1]:
                        filial = filial_normal[ind]
                full_fio = family.strip() + " " + name + otch
                # print(f'{filial} {num[0].value} {name}{otch} {num[2].value}')
                subscribers.append((filial, full_fio, num[2].value.strip()))
    print(subscribers)


def export_yealink():
    root = etree.Element("YealinkIPPhoneBook")
    title = etree.SubElement(root, "Title")
    title.text = "Yealink"
    deps = []
    for subscriber in subscribers:
        if subscriber[0] not in deps:
            deps.append(subscriber[0])
            menu = etree.SubElement(root, "Menu", Name=subscriber[0].replace('\n', ""))
            for i in subscribers:
                if i[0] == subscriber[0]:
                    etree.SubElement(menu, "Item", Name=i[1], Phone1=i[2].replace("-", ""))

    et = etree.ElementTree(root)
    et.write(output_file, pretty_print=True, encoding='utf-8', xml_declaration=True)


def export_eltex():
    root = etree.Element('EltexIPPhoneDirectory')
    title = etree.SubElement(root, 'Title')
    title.text = 'EltexPhones'
    prompt = etree.SubElement(root, 'Prompt')
    prompt.text = 'Prompt'
    gr_list = etree.SubElement(root, 'Grouplist')
    deps = []
    for subscriber in subscribers:
        if subscriber[0] not in deps:
            deps.append(subscriber[0])
            etree.SubElement(gr_list, 'Group', name=subscriber[0].replace('\n', ''))
    for subscriber in subscribers:
        dir_entry = etree.SubElement(root, 'DirectoryEntry')
        name = etree.SubElement(dir_entry, "Name")
        name.text = subscriber[1]
        telephone = etree.SubElement(dir_entry, 'Telephone')
        telephone.text = subscriber[2].replace('-', '')
        grp = etree.SubElement(dir_entry, 'Group')
        grp.text = subscriber[0]
    et = etree.ElementTree(root)
    et.write('eltex.xml', pretty_print=True, encoding = 'utf-8', xml_declaration=True)


if __name__ == '__main__':
    get_departments_names()
    get_normal_names()
    get_filial_rows()
    get_subs()
    export_yealink()
    export_eltex()
