import openpyxl as op
from lxml import etree

filename = 'ауп.xlsx'
wb = op.load_workbook(filename)
sheet_ranges = wb['Телефоны АУП']
max_row = sheet_ranges.max_row

departments = ['Руководство ',
               'Первичная профсоюзная организация                                                                                             "Газпром трансгаз Югорск профсоюз"',
               'Служба по связям с общественностью и СМИ', 'Юридический отдел', 'Отдел кадров и трудовых отношений',
               'Отдел внутреннего аудита', 'Бухгалтерия', 'Отдел  налогов', 'Финансовый отдел',
               'Планово-экономический отдел', 'Отдел страхования', 'Отдел организации труда и заработной платы',
               'Отдел управления имуществом', 'Отдел подготовки и проведения закупок',
               ' Служба организации реконструкции и строительства основных фондов', ' Специальный отдел',
               'Служба корпоративной защиты', 'Отдел охраны окружающей среды и энергосбережения',
               'Медицинская служба', 'Отдел социального развития',
               'Специалисты  службы хозяйственного обеспечения УЭЗиС', 'Производственно-диспетчерская служба',
               'ПО по эксплуатации магистральных газопроводов и ГРС', 'ПО по эксплуатации компрессорных станций',
               'Отдел главного энергетика', 'Производственный отдел защиты от коррозии',
               'ПО по эксплуатации подземных хранилищ газа', 'Отдел главного механика', 'Отдел главного сварщика',
               'Технический отдел', 'Отдел охраны труда', 'Служба промышленной и пожарной безопасности',
               'Производственный отдел метрологического обеспечения', 'Производственный отдел автоматизации',
               'Специалисты Инженерно-технического центра. \nОтдел сопровождения СОДУ. Служба АиМО',
               'Специалисты Управления Связи.                                                                                                                         Технический отдел',
               'Служба информационно-управляющих систем', 'Транспортный отдел', 'Служба организации вахтовых перевозок',
               'Отдел документационного обеспечения управления',
               'Специалисты Управления по эксплуатации зданий и сооружений   \nпо обслуживанию АКЗ ООО "Газпром трансгаз Югорск"',
               'Специалисты Управления Связи по обслуживанию средств связи АКЗ ООО "Газпром трансгаз Югорск"\nЦентральная служба связи ',
               'Комбинат общественного питания']
departments_normal = ['Руководство ', 'Профсоюз', 'ССО и СМИ', 'Юридический отдел', 'ОКиТО',
                      'Отдел внутреннего аудита', 'Бухгалтерия', 'Отдел налогов', 'Финансовый отдел',
                      'Планово-экономический отдел', 'Отдел страхования', 'Отдел организации труда и заработной платы',
                      'Отдел управления имуществом', 'Отдел подготовки и проведения закупок',
                      'СОРИСОФ', 'Специальный отдел', 'СКЗ', 'Отдел ООСиЭ', 'Медицинская служба',
                      'Отдел социального развития',
                      'Специалисты  службы хозяйственного обеспечения УЭЗиС', 'ПДС',
                      'ПО по ЭМГ и ГРС', 'ПО по ЭКС',
                      'Отдел главного энергетика', 'ПО защиты от коррозии',
                      'ПО по ЭПХГ', 'Отдел главного механика', 'Отдел главного сварщика',
                      'Технический отдел', 'Отдел охраны труда', 'СПиПБ',
                      'ПО метрологического обеспечения', 'ПО автоматизации',
                      'Специалисты ИТЦ. Отдел сопровождения СОДУ. Служба АиМО',
                      'Специалисты УС Технический отдел',
                      'СИУС', 'Транспортный отдел', 'Служба организации вахтовых перевозок',
                      'Отдел документационного обеспечения управления',
                      'Специалисты УЭЗиС по обслуживанию АКЗ',
                      'Специалисты УС Центральная служба связи ',
                      'Комбинат общественного питания']
department_rows = []
department_ranges = []
subscribers = []


def get_deps_row():
    for num in sheet_ranges.iter_rows(min_row=1, max_col=1, max_row=max_row):
        for department in departments:
            if num[0].value == department:
                department_rows.append(num[0].row)
                # print(department)
    department_rows.append(max_row)
    for index, rng in enumerate(department_rows):
        if index + 1 < len(department_rows):
            department_ranges.append([rng, department_rows[index + 1] - 1])
    print(f'Найдено {len(department_ranges)} диапазонов для {len(departments)} филиалов')
    # print(department_rows)


def get_subs():
    filial = ""
    for num in sheet_ranges.iter_rows(min_row=9, max_col=3, max_row=max_row):
        # print(num[2].value)
        # 0 - Фамилия, 1 - должность, 2 - номер
        # print(f'{num[2].value}  {type(num[2].value)}')
        if num[2].value and num[0].value and (len(str(num[2].value).strip()) == 7 or
                                              ((len(str(num[2].value).strip()) == 6) and str(
                                                  num[2].value[2:3] == "-"))):
            # print(f'{num[2].row} {num[2].value}')
            if sheet_ranges.cell(row=num[0].row + 1, column=1).value and num[0].value.strip().find(" ") == -1:
                io_cell = sheet_ranges.cell(row=num[0].row + 1, column=1).value
                family = num[0].value
                name = io_cell[0:1] + "."
                otch = io_cell[io_cell.find(" ") + 1:io_cell.find(" ") + 2] + "."
                for ind, filial_range in enumerate(department_ranges):
                    if filial_range[0] < num[0].row < filial_range[1]:
                        filial = departments_normal[ind]
                full_fio = family.strip() + " " + name + otch
                # print(f'{filial} {num[0].value} {name}{otch} {num[2].value}')
                subscribers.append((filial, full_fio, num[2].value.strip()))
    # print(len(subscribers))


def export_phonebook_yealink():
    root = etree.Element("YealinkIPPhoneBook")
    title = etree.SubElement(root, "Title")
    title.text = "Yealink"
    deps = []
    for subscriber in subscribers:
        if subscriber[0] not in deps:
            deps.append(subscriber[0])
            menu = etree.SubElement(root, "Menu", Name=subscriber[0].replace('\n', ""))
            for i in subscribers:
                if i[0] == subscriber[0]:
                    etree.SubElement(menu, "Item", Name=i[1], Phone1=i[2].replace("-", ""))

    et = etree.ElementTree(root)
    et.write('output_aup.xml', pretty_print=True, encoding='utf-8', xml_declaration=True)


def export_phonebook_eltex():
    pass


if __name__ == '__main__':
    get_deps_row()
    get_subs()
    export_phonebook_yealink()
